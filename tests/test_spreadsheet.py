"""Testes de integridade e formatação da geração de XLSX."""

import io
from decimal import Decimal
from datetime import datetime
import openpyxl

from backend.app.models.entities import (
    Cliente, PlanoContas, Transacao, TipoMovimento, StatusRevisao, OrigemClassificacao
)
from backend.app.services.spreadsheet_generator import SpreadsheetGeneratorService


def test_generate_xlsx_bytes_and_structure():
    cliente = Cliente(nome="Cliente Teste Planilha", documento="30.000.000/0001-30")
    conta = PlanoContas(cliente=cliente, numero_conta="1.1.1.01", descricao="Banco Conta Movimento", tipo="Ativo")

    tx1 = Transacao(
        cliente=cliente,
        data=datetime(2026, 5, 10),
        descricao_banco="RECEBIMENTO CLIENTE XPTO",
        valor=Decimal("1500.00"),
        tipo_movimento=TipoMovimento.ENTRADA,
        conta_classificada=conta,
        origem_classificacao=OrigemClassificacao.REGRA_EXATA,
        confianca=Decimal("1.00"),
        status_revisao=StatusRevisao.CONFIRMADO
    )

    tx2 = Transacao(
        cliente=cliente,
        data=datetime(2026, 5, 11),
        descricao_banco="PAGAMENTO LUZ",
        valor=Decimal("320.45"),
        tipo_movimento=TipoMovimento.SAIDA,
        conta_classificada=None,
        origem_classificacao=None,
        confianca=None,
        status_revisao=StatusRevisao.PENDENTE
    )

    service = SpreadsheetGeneratorService()
    xlsx_bytes = service.generate_bytes([tx1, tx2], cliente=cliente)

    # Carregar bytes com openpyxl para verificar validade
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    assert ws.title == "Conciliação Bancária"
    # Linha 1 = Cabeçalho, Linhas 2 e 3 = Dados
    assert ws.max_row == 3
    assert ws.max_column == 9

    # Verificar cabeçalhos
    assert ws.cell(row=1, column=1).value == "Data"
    assert ws.cell(row=1, column=4).value == "Valor (R$)"

    # Verificar valores da primeira linha de dados
    assert ws.cell(row=2, column=1).value == "10/05/2026"
    assert ws.cell(row=2, column=2).value == "RECEBIMENTO CLIENTE XPTO"
    assert ws.cell(row=2, column=4).value == 1500.00
    assert ws.cell(row=2, column=5).value == "1.1.1.01"

    # Verificar linha pendente
    assert ws.cell(row=3, column=5).value == "NÃO CLASSIFICADO"
    assert ws.cell(row=3, column=9).value == "pendente"


def test_default_filename_generation():
    service = SpreadsheetGeneratorService()
    filename = service.get_default_filename("Empresa Exemplo S/A", datetime(2026, 6, 1))
    assert filename == "Empresa_Exemplo_S_A_Conciliacao_2026-06.xlsx"
