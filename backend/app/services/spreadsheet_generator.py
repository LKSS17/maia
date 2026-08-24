"""Módulo para geração padronizada de planilhas XLSX contábeis."""

import io
import re
from datetime import datetime
from decimal import Decimal
from typing import List, Optional
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.app.models.entities import Transacao, Cliente, TipoMovimento


class SpreadsheetGeneratorService:
    """Gera arquivos XLSX padronizados com formatação contábil e rastreabilidade."""

    def __init__(self):
        # Estilos padronizados
        self.font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        self.fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.fill_zebra = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
        self.border_thin = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

    def _sanitize_filename(self, name: str) -> str:
        return re.sub(r'[\\/*?:"<>| ]', "_", name.strip())

    def get_default_filename(self, cliente_nome: str, referencia: Optional[datetime] = None) -> str:
        """Gera nome padronizado no formato: Cliente_Conciliacao_AAAA-MM.xlsx."""
        ref = referencia or datetime.now()
        clean_name = self._sanitize_filename(cliente_nome)
        return f"{clean_name}_Conciliacao_{ref.strftime('%Y-%m')}.xlsx"

    def build_workbook(self, transacoes: List[Transacao], cliente: Optional[Cliente] = None) -> Workbook:
        """Constrói o objeto Workbook com todas as transações formatadas."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Conciliação Bancária"

        headers = [
            "Data",
            "Histórico Bancário",
            "Tipo",
            "Valor (R$)",
            "Nº Conta Contábil",
            "Descrição da Conta",
            "Origem Classificação",
            "Confiança",
            "Status Revisão"
        ]

        # Inserir cabeçalho
        ws.append(headers)

        for col_idx, col_name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = self.font_header
            cell.fill = self.fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.border_thin

        ws.row_dimensions[1].height = 28

        # Inserir dados
        for row_idx, tx in enumerate(transacoes, 2):
            num_conta = tx.conta_classificada.numero_conta if tx.conta_classificada else "NÃO CLASSIFICADO"
            desc_conta = tx.conta_classificada.descricao if tx.conta_classificada else "PENDENTE DE CLASSIFICAÇÃO"
            origem = tx.origem_classificacao.value if tx.origem_classificacao else "pendente"
            confianca_fmt = f"{float(tx.confianca) * 100:.0f}%" if tx.confianca is not None else "-"

            row_data = [
                tx.data.strftime("%d/%m/%Y"),
                tx.descricao_banco,
                tx.tipo_movimento.value,
                float(tx.valor),
                num_conta,
                desc_conta,
                origem,
                confianca_fmt,
                tx.status_revisao.value
            ]
            ws.append(row_data)

            # Estilização das células de dados
            fill = self.fill_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None)

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = self.border_thin
                if fill.fill_type:
                    cell.fill = fill

                # Alinhamentos e Formatos específicos
                if col_idx == 1:  # Data
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx == 3:  # Tipo
                    cell.alignment = Alignment(horizontal="center")
                    if tx.tipo_movimento == TipoMovimento.ENTRADA:
                        cell.font = Font(color="006100", bold=True)
                    else:
                        cell.font = Font(color="9C0006")
                elif col_idx == 4:  # Valor
                    cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in [5, 7, 8, 9]:
                    cell.alignment = Alignment(horizontal="center")

        # Ajuste automático de largura de colunas
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Habilitar filtro automático
        ws.auto_filter.ref = ws.dimensions

        return wb

    def generate_bytes(self, transacoes: List[Transacao], cliente: Optional[Cliente] = None) -> bytes:
        """Exporta o arquivo Excel diretamente como array de bytes (para nuvem ou download)."""
        wb = self.build_workbook(transacoes, cliente)
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def generate_file(self, transacoes: List[Transacao], output_path: str, cliente: Optional[Cliente] = None) -> str:
        """Salva a planilha diretamente no sistema de arquivos local."""
        wb = self.build_workbook(transacoes, cliente)
        dest_path = Path(output_path)
        dest_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(dest_path)
        return str(dest_path)
